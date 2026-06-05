"""Tests for the formatted-Excel export used by the download button.

Uses fully synthetic data — the exporter only formats; it never depends on real
invoice values.
"""

import io

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from mydata.excel_export import dataframe_to_xlsx

BASE_COLS = ['Επιχείρηση', 'Αρχείο', 'MARK', 'Σειρά', 'Α/Α', 'Τύπος', 'Ημερομηνία',
             'Γραμμή', 'Κωδικός', 'Περιγραφή', 'Ποσότητα', 'Καθαρή Αξία', 'ΦΠΑ', 'Σύνολο']
COST_COLS = ['Τιμή Κτήσης (χ Ποσότητα)', 'Καθαρό Κέρδος']
COLUMNS = BASE_COLS + COST_COLS

FAKE_MARK = '123456789012345'   # synthetic 15-digit id


def _df():
    rows = [
        # a sale (matched): cost 1.00, profit 99.00
        ['ACME', 'a.pdf', FAKE_MARK, 'SER', '1', '1.1', '2026-01-02',
         1, 'SKU.001', 'Widget', 10, 100.00, 24.00, 124.00, 1.00, 99.00],
        # a credit note (matched, negated): cost -9.80, profit -40.20
        ['ACME', 'b.pdf', '123456789012999', 'CRD', '2', '5.1', '2026-01-03',
         1, 'SKU.002', 'Returned widget', -4, -50.00, -12.00, -62.00, -9.80, -40.20],
    ]
    return pd.DataFrame(rows, columns=COLUMNS)


def _load(xlsx_bytes):
    return load_workbook(io.BytesIO(xlsx_bytes))


def _data_marks(ws):
    col = COLUMNS.index('MARK') + 1
    return {ws.cell(r, col).value for r in range(2, ws.max_row + 1)
            if ws.cell(r, col).value not in (None, '')}      # totals row MARK is blank


# --- base formatting (asserted on the first 'Όλα' sheet) ---------------------

def test_returns_nonempty_bytes():
    data = dataframe_to_xlsx(_df())
    assert isinstance(data, (bytes, bytearray)) and len(data) > 0


def test_structure_table_and_freeze():
    ws = _load(dataframe_to_xlsx(_df())).active
    assert ws.title == "Όλα"
    assert "Inv_All" in ws.tables
    assert ws.freeze_panes == "A2"
    assert [ws.cell(1, c).value for c in range(1, len(COLUMNS) + 1)] == COLUMNS


def test_mark_and_type_kept_as_text():
    ws = _load(dataframe_to_xlsx(_df())).active
    mark = ws.cell(2, COLUMNS.index('MARK') + 1)
    typ = ws.cell(2, COLUMNS.index('Τύπος') + 1)
    assert mark.value == FAKE_MARK and isinstance(mark.value, str) and mark.number_format == '@'
    assert typ.value == '1.1' and typ.number_format == '@'


def test_dates_are_real_dates():
    import datetime
    ws = _load(dataframe_to_xlsx(_df())).active
    cell = ws.cell(2, COLUMNS.index('Ημερομηνία') + 1)
    assert isinstance(cell.value, (datetime.date, datetime.datetime))
    assert cell.number_format == 'DD/MM/YYYY'


def test_money_format_and_negative_credit_note():
    ws = _load(dataframe_to_xlsx(_df())).active
    net_col = COLUMNS.index('Καθαρή Αξία') + 1
    assert '€' in ws.cell(2, net_col).number_format
    assert ws.cell(3, net_col).value == -50.00          # credit note stays negative


def test_line_cost_is_money_and_summed():
    ws = _load(dataframe_to_xlsx(_df())).active
    tr = ws.max_row
    cc = COLUMNS.index('Τιμή Κτήσης (χ Ποσότητα)') + 1
    assert ws.cell(2, cc).value == 1.00 and '€' in ws.cell(2, cc).number_format
    L = get_column_letter(cc)
    assert ws.cell(tr, cc).value == f"=SUM({L}2:{L}3)"   # total cost of goods


def test_profit_is_money_and_summed_in_totals():
    ws = _load(dataframe_to_xlsx(_df())).active
    tr = ws.max_row
    pc = COLUMNS.index('Καθαρό Κέρδος') + 1
    assert '€' in ws.cell(2, pc).number_format
    assert ws.cell(3, pc).value == -40.20
    L = get_column_letter(pc)
    assert ws.cell(tr, pc).value == f"=SUM({L}2:{L}3)"


# --- three-sheet split --------------------------------------------------------

def _split_df():
    rows = [
        ['ACME', 'a.pdf', '111111111111111', 'SER', '1', '1.1', '2026-01-02',
         1, 'SKU.001', 'W', 10, 100.00, 24.00, 124.00, 1.00, 99.00],          # matched
        ['ACME', 'b.pdf', '222222222222222', 'SER', '2', '1.1', '2026-01-03',
         1, 'SKU.002', 'X', 5, 50.00, 12.00, 62.00, 0.50, 49.50],             # matched
        ['ACME', 'c.pdf', '333333333333333', 'SER', '3', '1.1', '2026-01-04',
         1, 'SKU.999', 'Y', 2, 20.00, 4.00, 24.00, None, None],               # unmatched
    ]
    return pd.DataFrame(rows, columns=COLUMNS)


def test_three_sheets_when_cost_present():
    wb = _load(dataframe_to_xlsx(_split_df()))
    assert wb.sheetnames == ["Όλα", "Με Τιμή Κτήσης", "Χωρίς Τιμή Κτήσης"]
    assert _data_marks(wb["Όλα"]) == {'111111111111111', '222222222222222', '333333333333333'}
    assert _data_marks(wb["Με Τιμή Κτήσης"]) == {'111111111111111', '222222222222222'}
    assert _data_marks(wb["Χωρίς Τιμή Κτήσης"]) == {'333333333333333'}


def test_single_sheet_when_no_cost_columns():
    df = pd.DataFrame([['ACME', 'a.pdf', FAKE_MARK, 'SER', '1', '1.1', '2026-01-02',
                        1, 'SKU.001', 'W', 10, 100.00, 24.00, 124.00]], columns=BASE_COLS)
    wb = _load(dataframe_to_xlsx(df))
    assert wb.sheetnames == ["Τιμολόγια"]


def test_empty_dataframe_does_not_crash():
    wb = _load(dataframe_to_xlsx(pd.DataFrame(columns=COLUMNS)))
    # cost columns present -> still three sheets, each with just the header
    assert wb.sheetnames == ["Όλα", "Με Τιμή Κτήσης", "Χωρίς Τιμή Κτήσης"]
    ws = wb["Όλα"]
    assert [ws.cell(1, c).value for c in range(1, len(COLUMNS) + 1)] == COLUMNS

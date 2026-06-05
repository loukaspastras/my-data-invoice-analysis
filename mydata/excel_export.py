"""Render the invoice DataFrame to a properly formatted .xlsx (bytes).

Used by the Streamlit download button so the user gets a ready-to-use Excel —
styled table, frozen header, MARK/Τύπος kept as text, € money format with red
negatives (credit notes), real dates, and a totals row that nets out.

When cost data is present (column "Τιμή Κτήσης (χ Ποσότητα)"), the workbook is
split into three sheets:
    Όλα               — every line
    Με Τιμή Κτήσης     — only lines that matched a cost
    Χωρίς Τιμή Κτήσης  — only lines whose cost was not found
"""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Columns that must stay textual (IDs / codes — never numbers or dates).
TEXT_COLS = ("MARK", "Σειρά", "Α/Α", "Τύπος", "Κωδικός")
DATE_COLS = ("Ημερομηνία",)
# Money columns: 2-decimal €, right-aligned, summed in the totals row.
MONEY_COLS = ("Καθαρή Αξία", "ΦΠΑ", "Σύνολο", "Τιμή Κτήσης (χ Ποσότητα)", "Καθαρό Κέρδος")
QTY_COLS = ("Ποσότητα",)
CENTER_INT_COLS = ("Γραμμή",)

# Presence of this column means cost data is available -> split into 3 sheets.
LINE_COST_COL = "Τιμή Κτήσης (χ Ποσότητα)"

MONEY_FMT = '#,##0.00" €";[Red]-#,##0.00" €"'
QTY_FMT = '#,##0.##;[Red]-#,##0.##'
WIDTH_CAPS = {"Περιγραφή": 55, "Αρχείο": 40}


def _write_sheet(ws, df, table_name):
    """Render one DataFrame into a worksheet with full formatting."""
    df = df.copy()
    cols = list(df.columns)
    idx = {c: i + 1 for i, c in enumerate(cols)}
    n = len(df)

    for c in TEXT_COLS:
        if c in df.columns:
            df[c] = df[c].apply(lambda v: "" if pd.isna(v) else str(v))
    parsed_dates = {c: pd.to_datetime(df[c], errors="coerce") for c in DATE_COLS if c in df.columns}

    ws.append(cols)
    for ri in range(n):
        row = []
        for c in cols:
            if c in parsed_dates:
                d = parsed_dates[c].iloc[ri]
                row.append(None if pd.isna(d) else d.date())
            else:
                v = df[c].iloc[ri]
                row.append(None if (not isinstance(v, str) and pd.isna(v)) else v)
        ws.append(row)

    def cells(col):
        L = get_column_letter(idx[col])
        return [ws[f"{L}{r}"] for r in range(2, n + 2)]

    for c in MONEY_COLS:
        if c in idx:
            for cell in cells(c):
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
    for c in QTY_COLS:
        if c in idx:
            for cell in cells(c):
                cell.number_format = QTY_FMT
                cell.alignment = Alignment(horizontal="right")
    for c in CENTER_INT_COLS:
        if c in idx:
            for cell in cells(c):
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
    for c in DATE_COLS:
        if c in idx:
            for cell in cells(c):
                cell.number_format = "DD/MM/YYYY"
    for c in TEXT_COLS:
        if c in idx:
            for cell in cells(c):
                cell.number_format = "@"

    # Totals row (nets out: credit notes are negative).
    present_money = [c for c in MONEY_COLS if c in idx]
    if n and present_money:
        tr = n + 2
        label_col = idx.get("Ποσότητα") or idx.get("Περιγραφή") or 1
        ws.cell(tr, label_col, "ΣΥΝΟΛΟ").alignment = Alignment(horizontal="right")
        for c in present_money:
            L = get_column_letter(idx[c])
            ws.cell(tr, idx[c], f"=SUM({L}2:{L}{n + 1})").number_format = MONEY_FMT
        fill = PatternFill("solid", fgColor="DDEBF7")
        top = Border(top=Side(style="double"))
        for col in range(1, len(cols) + 1):
            cc = ws.cell(tr, col)
            cc.font = Font(bold=True)
            cc.border = top
            cc.fill = fill

    # Styled, filterable table over header + data (excluding totals row).
    if n:
        ref = f"A1:{get_column_letter(len(cols))}{n + 1}"
        tbl = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True,
                                            showFirstColumn=False, showLastColumn=False,
                                            showColumnStripes=False)
        ws.add_table(tbl)

    for col in range(1, len(cols) + 1):
        ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for c in cols:
        L = get_column_letter(idx[c])
        longest = max([len(str(c))] + [len(str(v)) for v in df[c].astype(str).tolist()] + [4])
        ws.column_dimensions[L].width = min(longest + 2, WIDTH_CAPS.get(c, 22))

    ws.freeze_panes = "A2"


def dataframe_to_xlsx(df, sheet_name="Τιμολόγια"):
    """Return raw .xlsx bytes. Splits into 3 sheets when cost data is present."""
    wb = Workbook()
    if LINE_COST_COL in df.columns:
        matched = df[df[LINE_COST_COL].notna()]
        unmatched = df[df[LINE_COST_COL].isna()]
        ws_all = wb.active
        ws_all.title = "Όλα"
        _write_sheet(ws_all, df, "Inv_All")
        _write_sheet(wb.create_sheet("Με Τιμή Κτήσης"), matched, "Inv_Matched")
        _write_sheet(wb.create_sheet("Χωρίς Τιμή Κτήσης"), unmatched, "Inv_Unmatched")
    else:
        ws = wb.active
        ws.title = sheet_name
        _write_sheet(ws, df, "Invoices")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
